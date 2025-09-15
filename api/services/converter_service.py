"""
File conversion service for Excel/CSV to JSON - Azure Functions version
"""

import pandas as pd
import io
import numpy as np
import math
import json
import time
from typing import Dict, Any, List, Optional, Union

from models.requests import ConversionRequest
from services.ai_service import AIService
from utils.logger import logger
from utils.exceptions import (
    FileProcessingError, 
    UnsupportedFileFormatError,
    ValidationError
)
from core.config import get_azure_settings


class ConverterService:
    """Service for file conversion operations - Azure Functions version"""
    
    def __init__(self):
        self.ai_service = AIService()
        self.settings = get_azure_settings()
        self.max_file_size = self.settings.max_file_size
        self.allowed_extensions = self.settings.allowed_extensions
        logger.info("Converter Service initialized for Azure Functions")
    
    def _clean_dataframe(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Clean DataFrame from problematic values that can't be JSON serialized
        
        Args:
            df: Input DataFrame
            
        Returns:
            Cleaned DataFrame
        """
        df_clean = df.copy()
        
        # Replace inf and -inf with None
        df_clean = df_clean.replace([np.inf, -np.inf], None)
        
        # Convert any numpy types to native Python types
        for col in df_clean.columns:
            if df_clean[col].dtype == 'object':
                df_clean[col] = df_clean[col].apply(self._clean_single_value)
            elif pd.api.types.is_numeric_dtype(df_clean[col]):
                df_clean[col] = df_clean[col].apply(self._clean_numeric_value)
        
        return df_clean
    
    def _clean_single_value(self, value):
        """Clean individual value for JSON serialization"""
        if pd.isna(value):
            return None
        elif isinstance(value, (np.integer, np.floating)):
            if pd.isna(value) or np.isinf(value):
                return None
            return value.item()
        else:
            return value
    
    def _clean_numeric_value(self, value):
        """Clean numeric values"""
        if pd.isna(value) or np.isinf(value):
            return None
        return value
    
    async def convert_excel_to_json(
        self,
        file_content: bytes,
        filename: str,
        use_ai: bool = True,
        min_confidence: float = 0.8,
        sheet_name: Optional[str] = None,
        skip_rows: int = 0,
        max_rows: Optional[int] = None
    ) -> Dict[str, Any]:
        """
        Convert Excel/CSV file to JSON format - Azure Functions version
        
        Args:
            file_content: File content as bytes
            filename: Name of the file
            use_ai: Whether to use AI analysis
            min_confidence: Minimum confidence threshold for AI
            sheet_name: Specific sheet for Excel files
            skip_rows: Number of rows to skip
            max_rows: Maximum number of rows to process
            
        Returns:
            Conversion result with data and metadata
        """
        
        start_time = time.time()
        
        try:
            logger.info(f"🔋 Processing file: {filename}")
            
            # Validate file
            self._validate_file_bytes(file_content, filename)
            
            # Process file based on type
            df = self._process_file_bytes(filename, file_content, sheet_name, skip_rows, max_rows)
            
            # Clean DataFrame
            df = self._clean_dataframe(df)
            
            # AI analysis if requested
            ai_analysis = {}
            ai_usage = {}
            
            if use_ai:
                try:
                    ai_analysis, ai_usage = await self.ai_service.analyze_dataframe(
                        df, 
                        use_ai,
                        min_confidence,
                        filename
                    )
                except Exception as e:
                    logger.warning(f"AI analysis failed: {str(e)}")
                    ai_analysis = {"error": str(e)}
            
            # Convert to JSON
            json_data = self._dataframe_to_json(df)
            
            # Create metadata with all required fields
            processing_time = time.time() - start_time
            
            # Create default AI analysis if none provided
            if not ai_analysis:
                ai_analysis = {
                    "confidence": 95.0,
                    "analysis_type": "deterministic",
                    "detected_patterns": [],
                    "column_types": {col: str(df[col].dtype) for col in df.columns},
                    "recommendations": []
                }
            
            # Create default AI usage if none provided
            if not ai_usage:
                ai_usage = {
                    "ai_used": use_ai,
                    "processing_mode": "deterministic" if not use_ai else "ai_assisted",
                    "trigger_reason": None,
                    "issues_detected": [],
                    "ai_improvements": [],
                    "user_friendly_explanation": "File processed successfully using standard analysis.",
                    "technical_details": None
                }
            
            metadata = {
                "record_count": len(json_data),
                "columns": list(df.columns),
                "ai_analysis": ai_analysis,
                "ai_usage": ai_usage,
                "confidence": ai_analysis.get("confidence", 95.0) if isinstance(ai_analysis, dict) else getattr(ai_analysis, 'confidence', 95.0),
                "processing_time": f"{processing_time:.2f}s",
                "file_info": {
                    "filename": filename,
                    "original_rows": len(df),
                    "file_size_mb": round(len(file_content) / (1024 * 1024), 2),
                    "sheet_name": sheet_name or "Default" if filename.lower().endswith(('.xlsx', '.xls')) else None
                }
            }

            logger.info(f" File processed successfully in {processing_time:.2f}s")
            
            return {
                "status": "SUCCESS",
                "data": json_data,
                "metadata": metadata
            }
            
        except Exception as e:
            logger.error(f"ERROR: Conversion failed: {str(e)}")
            raise FileProcessingError(f"Conversion failed: {str(e)}")
    
    def _validate_file_bytes(self, file_content: bytes, filename: str) -> None:
        """
        Validate file content and format for Azure Functions
        """
        
        # Check file size
        file_size = len(file_content)
        if file_size > self.max_file_size:
            raise ValidationError(
                f"File size ({file_size} bytes) exceeds maximum allowed size ({self.max_file_size} bytes)"
            )
        
        # Check file extension
        file_extension = None
        for ext in self.allowed_extensions:
            if filename.lower().endswith(ext):
                file_extension = ext
                break
        
        if not file_extension:
            raise UnsupportedFileFormatError(
                f"File format not supported. Allowed formats: {', '.join(self.allowed_extensions)}"
            )
        
        logger.info(f"INFO: File validation passed: {filename} ({file_size} bytes)")
    
    def _process_file_bytes(
        self, 
        filename: str, 
        contents: bytes, 
        sheet_name: Optional[str] = None,
        skip_rows: int = 0,
        max_rows: Optional[int] = None
    ) -> pd.DataFrame:
        """
        Process file content based on file type - Azure Functions version
        """
        
        if filename.lower().endswith('.csv'):
            return self._process_csv(contents, skip_rows, max_rows)
        elif filename.lower().endswith(('.xlsx', '.xls')):
            return self._process_excel(contents, sheet_name, skip_rows, max_rows)
        else:
            raise UnsupportedFileFormatError(f"Unsupported file format: {filename}")
    
    def _process_csv(self, contents: bytes, skip_rows: int = 0, max_rows: Optional[int] = None) -> pd.DataFrame:
        """Process CSV file from bytes"""
        try:
            csv_io = io.StringIO(contents.decode('utf-8'))
            df = pd.read_csv(csv_io, skiprows=skip_rows, nrows=max_rows)
            logger.info(f"INFO: CSV data loaded: {len(df)} rows, {len(df.columns)} columns")
            return df
        except Exception as e:
            logger.error(f"CSV processing failed: {str(e)}")
            raise FileProcessingError(f"Failed to process CSV file: {str(e)}")
    
    def _process_excel(
        self, 
        contents: bytes, 
        sheet_name: Optional[str] = None,
        skip_rows: int = 0,
        max_rows: Optional[int] = None
    ) -> pd.DataFrame:
        """Process Excel file from bytes"""
        try:
            excel_io = io.BytesIO(contents)
            
            if sheet_name:
                df = pd.read_excel(excel_io, sheet_name=sheet_name, skiprows=skip_rows, nrows=max_rows)
            else:
                df = pd.read_excel(excel_io, skiprows=skip_rows, nrows=max_rows)
            
            logger.info(f"INFO: Excel data loaded: {len(df)} rows, {len(df.columns)} columns")
            return df
        except Exception as e:
            logger.error(f"Excel processing failed: {str(e)}")
            raise FileProcessingError(f"Failed to process Excel file: {str(e)}")
    
    def _dataframe_to_json(self, df: pd.DataFrame) -> List[Dict[str, Any]]:
        """
        Convert DataFrame to JSON format
        """
        try:
            # Convert to dict records
            json_data = df.to_dict('records')
            
            # Clean the data for JSON serialization
            cleaned_data = []
            for record in json_data:
                cleaned_record = {}
                for key, value in record.items():
                    cleaned_record[key] = self._clean_single_value(value)
                cleaned_data.append(cleaned_record)
            
            return cleaned_data
            
        except Exception as e:
            logger.error(f"DataFrame to JSON conversion failed: {str(e)}")
            raise FileProcessingError(f"Failed to convert data to JSON: {str(e)}")
    
    async def convert_json_to_excel(
        self, 
        json_data: Union[List[Dict], Dict], 
        filename: str = "converted_data.xlsx",
        apply_formatting: bool = True
    ) -> Dict[str, Any]:
        """
        Convert JSON data to Excel format - Azure Functions version
        """
        
        try:
            logger.info(f"🔄 Converting JSON to Excel: {filename}")
            
            # Convert JSON to DataFrame
            df = self._json_to_dataframe(json_data)
            
            # Clean DataFrame
            df = self._clean_dataframe(df)
            
            # Create Excel file in memory
            excel_buffer = io.BytesIO()
            
            with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Data', index=False)
                
                if apply_formatting:
                    # Apply basic formatting
                    worksheet = writer.sheets['Data']
                    
                    # Header formatting
                    from openpyxl.styles import Font, PatternFill
                    header_font = Font(bold=True)
                    header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                    
                    for col_num, value in enumerate(df.columns, 1):
                        cell = worksheet.cell(row=1, column=col_num)
                        cell.font = header_font
                        cell.fill = header_fill
                    
                    # Auto-adjust column widths
                    for column in worksheet.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)
                        worksheet.column_dimensions[column_letter].width = adjusted_width
            
            excel_content = excel_buffer.getvalue()
            
            metadata = {
                "filename": filename,
                "rows": len(df),
                "columns": len(df.columns),
                "file_size_kb": round(len(excel_content) / 1024, 2),
                "formatting_applied": apply_formatting
            }
            
            logger.info(f"✅ JSON to Excel conversion completed: {len(df)} rows")
            
            return {
                "excel_content": excel_content,
                "filename": filename,
                "metadata": metadata
            }
            
        except Exception as e:
            logger.error(f"ERROR: JSON to Excel conversion failed: {str(e)}")
            raise FileProcessingError(f"JSON to Excel conversion failed: {str(e)}")
    
    def _json_to_dataframe(self, json_data: Union[List[Dict], Dict]) -> pd.DataFrame:
        """Convert JSON data to DataFrame"""
        try:
            if isinstance(json_data, dict):
                # If single object, convert to list
                json_data = [json_data]
            
            df = pd.DataFrame(json_data)
            return df
            
        except Exception as e:
            logger.error(f"JSON to DataFrame conversion failed: {str(e)}")
            raise FileProcessingError(f"Failed to convert JSON to DataFrame: {str(e)}")
    
    async def convert_excel_to_sql(
        self,
        file_content: bytes,
        filename: str,
        table_name: str = "converted_data",
        include_create_table: bool = True,
        include_inserts: bool = True
    ) -> Dict[str, Any]:
        """
        Convert Excel file to SQL queries - Azure Functions version
        """
        
        try:
            logger.info(f"INFO: Converting Excel to SQL: {filename} -> {table_name}")
            
            # Validate file
            self._validate_file_bytes(file_content, filename)
            
            # Process Excel file
            df = self._process_excel(file_content)
            
            # Clean data
            df = self._clean_dataframe(df)
            
            sql_queries = {}
            
            if include_create_table:
                # Generate CREATE TABLE statement
                create_table = self._generate_create_table_sql(df, table_name)
                sql_queries["create_table"] = create_table
            
            if include_inserts:
                # Generate INSERT statements
                insert_statements = self._generate_insert_sql(df, table_name)
                sql_queries["insert_statements"] = insert_statements
            
            metadata = {
                "table_name": table_name,
                "rows": len(df),
                "columns": len(df.columns),
                "create_table_included": include_create_table,
                "inserts_included": include_inserts,
                "total_statements": (1 if include_create_table else 0) + (len(insert_statements) if include_inserts else 0)
            }
            
            logger.info(f"✅ Excel to SQL conversion completed: {metadata['total_statements']} statements")
            
            return {
                "sql_queries": sql_queries,
                "metadata": metadata
            }
            
        except Exception as e:
            logger.error(f"ERROR: Excel to SQL conversion failed: {str(e)}")
            raise FileProcessingError(f"Excel to SQL conversion failed: {str(e)}")
    
    def _generate_create_table_sql(self, df: pd.DataFrame, table_name: str) -> str:
        """Generate CREATE TABLE SQL statement from DataFrame"""
        
        columns = []
        for col_name, dtype in df.dtypes.items():
            # Sanitize column name
            safe_col_name = col_name.replace(' ', '_').replace('-', '_')
            
            # Map pandas dtype to SQL type
            if pd.api.types.is_integer_dtype(dtype):
                sql_type = "INT"
            elif pd.api.types.is_float_dtype(dtype):
                sql_type = "DECIMAL(18,2)"
            elif pd.api.types.is_datetime64_any_dtype(dtype):
                sql_type = "DATETIME"
            else:
                sql_type = "NVARCHAR(255)"
            
            columns.append(f"    [{safe_col_name}] {sql_type}")
        
        create_table_sql = f"""CREATE TABLE [{table_name}] (
{chr(10).join(columns)}
);"""
        
        return create_table_sql
    
    def _generate_insert_sql(self, df: pd.DataFrame, table_name: str) -> List[str]:
        """Generate INSERT SQL statements from DataFrame"""
        
        # Sanitize column names
        safe_columns = [col.replace(' ', '_').replace('-', '_') for col in df.columns]
        columns_str = ', '.join([f'[{col}]' for col in safe_columns])
        
        insert_statements = []
        
        for _, row in df.iterrows():
            values = []
            for val in row:
                if pd.isna(val):
                    values.append("NULL")
                elif isinstance(val, str):
                    escaped_val = val.replace("'", "''")
                    values.append(f"'{escaped_val}'")
                else:
                    values.append(str(val))
            
            values_str = ', '.join(values)
            insert_sql = f"INSERT INTO [{table_name}] ({columns_str}) VALUES ({values_str});"
            insert_statements.append(insert_sql)
        
        return insert_statements
